"""
Excel Import/Export utilities for all modules.
"""

from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_to_excel(queryset, columns, filename='export'):
    """
    Export a queryset to an Excel file.

    columns: list of dicts with keys:
        - 'field': model field name or callable
        - 'header': Arabic header text
        - 'width': column width (optional, default 15)
        - 'format': number format (optional)
    """
    wb = Workbook()
    try:
        ws = wb.active
        ws.sheet_view.rightToLeft = True

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col['header'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = col.get('width', 15)

        for row_idx, obj in enumerate(queryset, 2):
            for col_idx, col in enumerate(columns, 1):
                field = col['field']
                if callable(field):
                    value = field(obj)
                else:
                    value = getattr(obj, field, '')
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%d/%m/%Y')
                    elif hasattr(value, '__str__') and not isinstance(value, int | float | str):
                        value = str(value)

                if isinstance(value, str) and value and value[0] in '=+-@':
                    value = "'" + value

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                if 'format' in col:
                    cell.number_format = col['format']

        ws.auto_filter.ref = ws.dimensions

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename="{filename}_{ts}.xlsx"'
        wb.save(response)
        return response
    finally:
        wb.close()


def import_from_excel(file, columns):
    """
    Import data from an uploaded Excel file.

    columns: list of dicts with keys:
        - 'field': model field name
        - 'header': Excel column header to match
        - 'type': 'str', 'int', 'float', 'date' (optional, default 'str')
        - 'required': bool (optional, default False)

    Returns: list of dicts with the parsed rows.
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    header_row = [str(cell.value or '').strip() for cell in ws[1]]
    col_map = {}
    for col_def in columns:
        header = col_def['header']
        for idx, h in enumerate(header_row):
            if h == header:
                col_map[idx] = col_def
                break

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        has_data = False
        for idx, col_def in col_map.items():
            if idx < len(row):
                value = row[idx]
            else:
                value = None

            field = col_def['field']
            col_type = col_def.get('type', 'str')

            if value is None or (isinstance(value, str) and value.strip() == ''):
                row_data[field] = None
                continue

            has_data = True

            if col_type == 'int':
                try:
                    row_data[field] = int(float(value))
                except (ValueError, TypeError):
                    row_data[field] = 0
            elif col_type == 'float':
                try:
                    row_data[field] = float(value)
                except (ValueError, TypeError):
                    row_data[field] = 0.0
            elif col_type == 'decimal':
                try:
                    row_data[field] = Decimal(str(value).replace(',', '').strip())
                except (ValueError, TypeError, InvalidOperation):
                    row_data[field] = None
            elif col_type == 'date':
                if isinstance(value, datetime):
                    row_data[field] = value.date()
                else:
                    try:
                        from datetime import datetime as dt

                        row_data[field] = dt.strptime(str(value)[:10], '%Y-%m-%d').date()
                    except Exception:
                        row_data[field] = None
            else:
                row_data[field] = str(value).strip() if value else ''

        if has_data and any(v is not None for v in row_data.values()):
            rows.append(row_data)

    wb.close()
    return rows
